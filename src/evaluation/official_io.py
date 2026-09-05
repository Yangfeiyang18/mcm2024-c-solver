from __future__ import annotations

import csv
import json
import shutil
import warnings
from collections import defaultdict
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[2]
CLEAN_DIR = PROJECT_ROOT / "数据" / "清洗后数据"
RESULTS_DIR = PROJECT_ROOT / "results"
FINAL_RESULTS_DIR = RESULTS_DIR / "final"
OUTPUTS_DIR = PROJECT_ROOT / "outputs"
FINAL_OUTPUTS_DIR = OUTPUTS_DIR / "final"
TEMPLATE_DIR = PROJECT_ROOT / "附件3"
FIGURES_DIR = PROJECT_ROOT / "figures" / "final"
AREA_TOLERANCE = 1e-6
DATA_ROWS = range(2, 84)
DATA_COLS = range(3, 44)
YEARS = list(range(2024, 2031))

OFFICIAL_BOOKS = {
    "result1_1.xlsx": RESULTS_DIR / "q1_full_waste_solution_long.csv",
    "result1_2.xlsx": RESULTS_DIR / "q1_full_discount_solution_long.csv",
    "result2.xlsx": RESULTS_DIR / "q2_lambda_025_solution_long.csv",
}

LOCKED = {
    "q1_waste_profit_yuan": 40768549.264908,
    "q1_discount_profit_yuan": 58465088.158657,
    "q1_waste_profit_wan": 4076.85,
    "q1_discount_profit_wan": 5846.51,
    "q1_waste_gap": 0.010373681922056817,
    "q1_discount_gap": 0.007700934077326965,
    "q2_training_mean_yuan": 40918998.912928,
    "q2_training_cvar_yuan": 40608592.438851,
    "q2_gap": 0.02509909801471907,
    "q2_test_independent_mean_yuan": 40921731.80517,
    "q2_test_independent_std_yuan": 178081.973686,
    "q2_test_independent_cvar_yuan": 40608434.188998,
    "q2_test_weak_mean_yuan": 40871491.505276,
    "q2_test_weak_std_yuan": 197274.166893,
    "q2_test_weak_cvar_yuan": 40519309.464777,
    "q2_test_medium_mean_yuan": 40859266.700947,
    "q2_test_medium_std_yuan": 232188.011891,
    "q2_test_medium_cvar_yuan": 40449257.410898,
    "obsolete_q1_waste_yuan": 40757749.0,
    "obsolete_q1_discount_yuan": 58380972.0,
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def load_mapping() -> dict[tuple[int, str, str, int], dict[str, Any]]:
    mapping: dict[tuple[int, str, str, int], dict[str, Any]] = {}
    for row in read_csv(CLEAN_DIR / "template_mapping.csv"):
        key = (int(row["year"]), row["season"], row["plot_id"], int(row["crop_id"]))
        mapping[key] = {
            "sheet": str(int(row["sheet"])),
            "row": int(row["row"]),
            "column": int(row["column"]),
            "cell": row["cell"],
            "crop_name": row["crop_name"],
            "template_season": row["template_season"],
        }
    return mapping


def load_area_by_key(path: Path) -> dict[tuple[int, str, str, int], float]:
    area: dict[tuple[int, str, str, int], float] = defaultdict(float)
    for row in read_csv(path):
        area[(int(row["year"]), row["season"], row["plot_id"], int(row["crop_id"]))] += float(row["area_mu"])
    return dict(area)


def expected_cells(
    area_by_key: dict[tuple[int, str, str, int], float],
    mapping: dict[tuple[int, str, str, int], dict[str, Any]],
) -> dict[tuple[str, int, int], float]:
    cells: dict[tuple[str, int, int], float] = defaultdict(float)
    missing: list[str] = []
    for key, area in area_by_key.items():
        if area <= AREA_TOLERANCE:
            continue
        location = mapping.get(key)
        if location is None:
            missing.append(str(key))
            continue
        cells[(location["sheet"], location["row"], location["column"])] += area
    if missing:
        raise KeyError(f"方案有 {len(missing)} 条记录找不到模板映射，例如 {missing[:5]}")
    return dict(cells)


def _load_workbook(path: Path):
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Cannot parse header or footer")
        return load_workbook(path)


def read_excel_data_cells(path: Path) -> dict[tuple[str, int, int], float]:
    workbook = _load_workbook(path)
    values: dict[tuple[str, int, int], float] = {}
    for year in YEARS:
        sheet = workbook[str(year)]
        for row in DATA_ROWS:
            for column in DATA_COLS:
                value = sheet.cell(row, column).value
                if value in (None, ""):
                    continue
                values[(str(year), row, column)] = float(value)
    return values


def fill_official_workbook(template_path: Path, solution_path: Path, output_path: Path) -> dict[str, Any]:
    mapping = load_mapping()
    area_by_key = load_area_by_key(solution_path)
    expected = expected_cells(area_by_key, mapping)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)
    workbook = _load_workbook(output_path)
    for (sheet_name, row, column), area in expected.items():
        cell = workbook[sheet_name].cell(row, column)
        cell.value = round(float(area), 6)
        cell.number_format = "0.######"
    workbook.save(output_path)
    return verify_excel_against_csv(output_path, solution_path, mapping)


def verify_excel_against_csv(
    excel_path: Path,
    solution_path: Path,
    mapping: dict[tuple[int, str, str, int], dict[str, Any]] | None = None,
) -> dict[str, Any]:
    mapping = mapping or load_mapping()
    expected = expected_cells(load_area_by_key(solution_path), mapping)
    actual = read_excel_data_cells(excel_path)
    missing_in_excel = sorted(set(expected) - set(actual))
    extra_in_excel = sorted(set(actual) - set(expected))
    mismatches = []
    for key in sorted(set(expected) & set(actual)):
        if abs(expected[key] - actual[key]) > AREA_TOLERANCE:
            mismatches.append({
                "sheet": key[0],
                "row": key[1],
                "column": key[2],
                "csv_mu": expected[key],
                "excel_mu": actual[key],
                "diff_mu": actual[key] - expected[key],
            })
    csv_total = sum(expected.values())
    excel_total = sum(actual.values())
    passed = not missing_in_excel and not extra_in_excel and not mismatches
    return {
        "excel": str(excel_path.relative_to(PROJECT_ROOT)),
        "csv": str(solution_path.relative_to(PROJECT_ROOT)),
        "status": "PASS" if passed else "FAIL",
        "csv_positive_cells": len(expected),
        "excel_positive_cells": len(actual),
        "csv_total_mu": round(csv_total, 6),
        "excel_total_mu": round(excel_total, 6),
        "total_diff_mu": round(excel_total - csv_total, 8),
        "missing_in_excel": [f"{sheet}!r{row}c{column}" for sheet, row, column in missing_in_excel[:20]],
        "extra_in_excel": [f"{sheet}!r{row}c{column}" for sheet, row, column in extra_in_excel[:20]],
        "mismatch_count": len(mismatches),
        "mismatches": mismatches[:20],
    }


def close_enough(actual: float, expected: float, rel: float = 1e-8, abs_tol: float = 1.0) -> bool:
    return abs(float(actual) - float(expected)) <= max(abs_tol, rel * abs(float(expected)))
